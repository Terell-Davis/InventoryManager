import xlsxwriter
import openpyxl
import os

StoreInv = './Inventory.xlsx'
StoreSheet = 'Store Inv.'

StoreCount = './InventoryCount.xlsx'
StoreCountSheet = 'Inventory Count'

path = "./Inventory/"

files = os.listdir(path)

invdic = {}
invcount = {}

productname = []
onhandcount = []
productupc = []
countedqt = []

wbinv = openpyxl.load_workbook(StoreInv)
wbcounts = openpyxl.load_workbook(StoreCount)

InvSheet = wbinv[StoreSheet]
CountSheet = wbcounts[StoreCountSheet]

# UPC
count = 0
for upc in InvSheet['B']:
    # Set Key of invdic
    if not count % 2 == 0:
        invdic[upc.value] = []
    count += 1

# Desc
count = 0
for desc in InvSheet['A']:
    # add to product name list
    if not count % 2 == 0:
        productname.append(desc.value)
    count += 1

# OnHand
count = 0
for onhand in InvSheet['E']:
    # add to onhandcount list
    if not count % 2 == 0:
        onhandcount.append(onhand.value)
    count += 1

i = 0
for upc in invdic:
    # add all values to invdic
    invdic[upc] = [productname[i], onhandcount[i]]
    i += 1
# ==========================

# adds upcs from count sheets
for countupc in CountSheet['A']:
    invcount[countupc.value] = []

# list to add the quantities
for counted in CountSheet['C']:
    countedqt.append(counted.value)

# adding quantity to the associated key within dic
j = 0
for data in invcount:
    invcount[data] = [countedqt[j]]
    j += 1

diff = invdic.keys() - invcount.keys()
#print(diff)

finaltest = {}
for onhandupc in invdic:
    for invcountupc in invcount:
        if onhandupc == invcountupc:
            finaltest[onhandupc] = invdic.get()
            invdic[onhandupc].append(invcount.get(onhandupc))

for test in invdic:
    invdic[test].append("Not Counted")
    invdic[test].append("Test")

workbook = xlsxwriter.Workbook('InventoryCount.xlsx')
worksheet = workbook.add_worksheet("Inventory Final")

none = "None"
row = 1
for description, onHand, itemcount in invdic.items():
    worksheet.write(row, 0, str(description))
    worksheet.write(row, 2, str(onHand))
    worksheet.write(row, 3, str(itemcount))
    row += 1

noitem ={}
for notcounted in diff:
    noitem[notcounted] = "Not Counted"






